import os
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from datetime import date

from time import sleep


#COMO FAZER O DRIVE ABRIR UM CHROME JÁ ABERTO?


options = webdriver.ChromeOptions() 

options.add_argument('--user-data-dir=C:\\Users\\Plugify\\AppData\\Local\\Google\\Chrome\\User Data')

options.add_argument('--profile-directory=Default')

#LEMBRAR DE ATUALIZAR O CHROME DRIVER NA PASTINHA
driver = webdriver.Chrome(executable_path="C:\\Users\\Plugify\\.spyder-py3\\chromedriver.exe",options=options)


wait = WebDriverWait(driver, 400) 


#Faremos um try-catch // o código vai tentar abrir um email do export report, caso falhe, vai ir no CAP
try:
    
    driver.get('https://mail.google.com/mail/u/0/#inbox')
    sleep(4)
    driver.find_element(By.XPATH, "//*[@class='bog']//*[contains(text(),'CAPWorkflow - Export Report')]").click()
    

except Exception:
    
    driver.get("https://vtex.capworkflow.com/Pages/Home.aspx")
    
    try:
        driver.find_element(By.XPATH, "//span[normalize-space()='CAP Workflow']").click()
        sleep(5)
        driver.find_element(By.XPATH, "//input[@id='PlaceHolderPageBody_signInButton']").click()
    except Exception:
        pass

    #Analytics
    driver.get("https://vtex.capworkflow.com/Pages/Report.aspx?params=UmVwb3J0SWQ9MTY=")

    sleep(3)
    #FAVORITES
    wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div/div/ul/li[2]/a/span"))).click() #FAVORITES
    
    #PURCHASE
    wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div/div/div/div[2]/div/div/div/div[2]/div/div[1]/div/div[1]"))).click() #PURCHASE REQUEST
    
    # 28/12/2022 -> date.today()
    data = str(date.today())
    data2 = "28/12/2022"
 
    i = 0
    while i<10:
        driver.find_element(By.XPATH, "/html/body/div/section/div/div[2]/div/div[2]/div/form/div/div/div[4]/div/input").send_keys(Keys.BACKSPACE)
        i = i+1
    driver.find_element(By.XPATH, "/html/body/div/section/div/div[2]/div/div[2]/div/form/div/div/div[4]/div/input").send_keys(data2)
    
    i = 0
    while i<10:
        driver.find_element(By.XPATH, "/html/body/div/section/div/div[2]/div/div[2]/div/form/div/div/div[5]/div/input").send_keys(Keys.BACKSPACE)
        i = i+1
    driver.find_element(By.XPATH, "/html/body/div/section/div/div[2]/div/div[2]/div/form/div/div/div[5]/div/input").send_keys(data)
    
    driver.find_element(By.XPATH, "//button[normalize-space()='Export']").click()
    sleep(1)

    #VAMOS ABRIR O GMAIL PARA BAIXAR
    driver.get('https://mail.google.com/mail/u/0/#inbox')
    
    
    #ESPERANDO O RELATÓRIO APARECER .....
    
    wait.until(EC.visibility_of_element_located((By.XPATH, "//*[@class='bog']//*[contains(text(),'CAPWorkflow - Export Report')]"))).click()
    
    

sleep(2)

try:
    driver.find_element(By.XPATH,"//a[contains(@href,'https://vtex.capworkflow.com/Services/CAPBinaryContents.svc')]").click()	 #CLICOU NO DOWNLOAD
    
    try: #SE ENTRAR PRA FAZER O LOGIN CHATO
        driver.find_element(By.XPATH, "/html/body/form/div[3]/div/div/div[2]/div/div/div[1]/div/button").click()
        driver.find_element(By.XPATH, "/html/body/form/div[3]/div/div/div[2]/div/input").click()
    except Exception:
        sleep(2)
except Exception:
    driver.find_element(By.XPATH, "//*[@class='ajT']").click() #CLICAR NOS 3 PONTINHOS
    driver.find_element(By.XPATH,"//a[contains(@href,'https://vtex.capworkflow.com/Services/CAPBinaryContents.svc')]").click()	 #CLICOU NO DOWNLOAD
    
    try:#SE ENTRAR PRA FAZER O LOGIN CHATO
        driver.find_element(By.XPATH, "/html/body/form/div[3]/div/div/div[2]/div/div/div[1]/div/button").click()
        driver.find_element(By.XPATH, "/html/body/form/div[3]/div/div/div[2]/div/input").click()
    except Exception:
        sleep(2)

#ARQUIVAR O EMAIL
driver.find_element(By.XPATH, "/html/body/div[7]/div[3]/div/div[2]/div[5]/div/div/div/div/div[1]/div[2]/div[1]/div/div[2]/div[1]/div").click()



#Função pra retornar o nome do arquivo CSV que fizemos download
path = "C:\\Users\\Plugify\\Downloads"

files = os.listdir(path)
paths = [os.path.join(path, basename) for basename in files]
CAMINHO = max(paths, key=os.path.getctime)


     

#FUNÇÃO TRANSFORMAR CSV PARA XLSX 


f = open(CAMINHO)

csv.register_dialect('colons', delimiter=';')

reader = csv.reader(f, dialect='colons')

wb = Workbook()
dest_filename = r'C:\\Users\\Plugify\\Desktop\\Purchase6V.xlsx'

ws = wb.worksheets[0]

for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        
        ws['%s%s'%(column_letter, (row_index + 1))].value = cell
        
wb.save(filename = dest_filename)






#FUNÇÃO PARA A 6ª VERSÃO CAP - Purchase Request

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

wb = load_workbook('C:\\Users\\Plugify\\Desktop\\Purchase6V.xlsx')
ws = wb.active

for row1 in range(ws.min_row, ws.max_row + 1): #ajusta o tamanho das linhas
    ws.row_dimensions[row1].height = 20
    

#VAMOS CONSERTAR AQUI AS NOVAS COLUNAS da V6

for row in range(2, len(ws['A']) + 1): 
    linha = str(row)
    if row < 53:
        ws['BW' + linha] = ws['BX' + linha].value

for row in range(2, len(ws['A']) + 1): 
    linha = str(row)
    if row > 53:
        ws['BY' + linha] = ws['BZ' + linha].value

BX = coordinate_from_string('BX1')
BY = coordinate_from_string('BY1')
col1 = column_index_from_string(BX[0]) # returna o numero correspondente da coluna da célula BX1
col2 = column_index_from_string(BY[0])

ws.delete_cols(col1,1)
ws.delete_cols(col2,1)




#Inserir as colunas que faltam
ws['ci' + '1'] = "Supplier - ALL"
ws['cj' + '1'] = "Estimated Cost USD"
ws['ck' + '1'] = "Initial Proposal USD"
ws['cl' + '1'] = "Final Proposal USD"
ws['cm' + '1'] = "Saving USD"
ws['cn' + '1'] = "'%' saving"
ws['co' + '1'] = "Requested ITEMS - fixed"


        
        
char2 = [] #TROCAR A VIRGULA POR PONTO SÓ PRA FAZER AS CONTAS AQUI NA FUNÇÃO PYTHON
for row in range(2, len(ws['A']) + 1): 
    char = str(row)
    
    if ws['S' + char].value != None: #EXCHANGE RATE
        char2 = str(ws['S' + char].value)
        char2 = char2.replace(",",".")
        ws['S' + char] = char2
        
    if ws['BO' + char].value != None: #INITIAL PROPOSAL
        char2 = str(ws['BO' + char].value)
        char2 = char2.replace(".","")
        char2 = char2.replace(",",".")
        ws['BO' + char] = char2
        
    if ws['BP' + char].value != None: #FINAL PROPOSAL
        char2 = str(ws['BP' + char].value)
        char2 = char2.replace(".","")
        char2 = char2.replace(",",".")
        ws['BP' + char] = char2
        
    if ws['BQ' + char].value != None: #SAVING
        char2 = str(ws['BQ' + char].value)
        char2 = char2.replace(".","")
        char2 = char2.replace(",",".")
        ws['BQ' + char] = char2
        
    if ws['CE' + char].value != None: #ESTIMATED TOTAL COST
        char2 = str(ws['CE' + char].value)
        char2 = char2.replace(".","")
        char2 = char2.replace(",",".")
        ws['CE' + char] = char2
        
    
        
for row in range(2,len(ws['A']) + 1): #PREENCHER A COLUNA SUPPLIER-ALL
    char = str(row)
    if(ws['BF' + char].value) != None:
        ws['CI' + char] = ws['BF' + char].value
    else:
        ws['CI' + char] = ws['AO' + char].value


for row in range(2,len(ws['A'])): #PREENCHER A COLUNA Estimated Total Cost USD
    char = str(row)
    ws['CJ' + char] = (ws['U' + char].value) #Coluna TOTAL USD 

        #########################################    
            
            
for row in range(2,len(ws['A']) + 1): #PREENCHER A COLUNA Initial Proposal
    char = str(row)

    try:
        ws['CK' + char] = (int(float(ws['BO' + char].value))/float(ws['S' + char].value))
            
    except Exception:
        ws['CK' + char] = '-'
        
for row in range(2,len(ws['A']) + 1): #PREENCHER A COLUNA Final Proposal
    char = str(row)

    try:
        ws['CL' + char] = (int(float(ws['BP' + char].value))/float(ws['S' + char].value))
            
    except Exception:
        ws['CL' + char] = '-'       

for row in range(2,len(ws['A']) + 1): #PREENCHER A COLUNA SAVING
    char = str(row)

    try:
        ws['CM' + char] = (int(float(ws['BQ' + char].value))/float(ws['S' + char].value))
            
    except Exception:
        ws['CM' + char] = '-'   


for row in range(2,len(ws['A']) + 1): #PREENCHER A COLUNA % SAVING
    char = str(row)

    try:
        ws['CN' + char] = (int(float(ws['CM' + char].value))/float(ws['CK' + char].value))
            
    except Exception:
        ws['CN' + char] = '-'   
        
        
char2 = []    #TROCAR ponto pela vírgula (Acabou as contas aí na parte de cima e agora podemos trocar de volta)
for row in range(2, len(ws['A']) + 1): 
    char = str(row)
    
    if ws['S' + char].value != None:
        char2 = str(ws['S' + char].value)
        char2 = char2.replace(".",",")
        ws['S' + char] = char2
        
        
    if ws['BO' + char].value != None:
        char2 = str(ws['BO' + char].value)
        char2 = char2.replace(".",",")
        ws['BO' + char] = char2        

    if ws['BP' + char].value != None:
        char2 = str(ws['BP' + char].value)
        char2 = char2.replace(".",",")
        ws['BP' + char] = char2

    if ws['BQ' + char].value != None:
        char2 = str(ws['BQ' + char].value)
        char2 = char2.replace(".",",")
        ws['BQ' + char] = char2

    if ws['CE' + char].value != None:
        char2 = str(ws['CE' + char].value)
        char2 = char2.replace(".",",")
        ws['CE' + char] = char2
        
    if ws['CK' + char].value != None:
        char2 = str(ws['CK' + char].value)
        char2 = char2.replace(".",",")
        ws['CK' + char] = char2

    if ws['CL' + char].value != None:
        char2 = str(ws['CL' + char].value)
        char2 = char2.replace(".",",")
        ws['CL' + char] = char2
        
    if ws['CM' + char].value != None:
        char2 = str(ws['CM' + char].value)
        char2 = char2.replace(".",",")
        ws['CM' + char] = char2       

    #NAO TEM PRA CN 
        
        
        
        
# COLUNA DESCRIPTION OF ITEMS      
for row in range(2, len(ws['A']) + 1): 
    linha = str(row)
    char = ws['T' + linha].value
    descricao = char.split(',')
    Description = descricao[0]
    Description = Description.replace("[","")
    Description = Description.replace("\"","")
    ws['CO' + linha] = Description


    
#tira as celulas de branco por "-"

for row in range(1, len(ws['A']) + 1):
    linha = str(row)
    for col in range(ws.min_column, ws.max_column + 1):
    
        coluna = get_column_letter(col)
        if(ws[coluna + linha].value) == "" or(ws[coluna + linha].value) == None:
            ws[coluna + linha] = "-"
    
#Dá o espaçamento pra não dar erro do #####
for col in range(ws.min_column, ws.max_column + 1):
    coluna = get_column_letter(col)
    ws.column_dimensions[coluna].width = 30

wb.save('C:\\Users\\Plugify\\Desktop\\Purchase6V.xlsx')












