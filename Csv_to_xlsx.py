import os
import glob
import csv
from openpyxl import Workbook #Type ´pip install openpyxl´ in the terminal
from openpyxl.utils import get_column_letter

diretorio = os.path.join(os.path.expanduser('~'), 'Downloads')  # find the download directory path

array = []
contador = 0
arquivos_csv = glob.glob(os.path.join(diretorio, '*.csv'))  # list the .csv files
for arquivo in arquivos_csv:
    print(arquivo + " - " + str(contador))# print the csv archive's paths
    array.append(arquivo)
    contador = contador + 1


try:
    num = int(input("\nType an integer value from above: \n"))

    print(f"You chose {array[num]}")
    
    #TRANSFORM CSV TO XLSX
    
    
    f = open(array[num])
    csv.register_dialect('colons', delimiter=';')

    reader = csv.reader(f, dialect='colons')

    wb = Workbook()
    dest_filename = diretorio

    ws = wb.worksheets[0]

    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
                
            ws['%s%s'%(column_letter, (row_index + 1))].value = cell
                
    wb.save(filename = dest_filename)

            
    
except ValueError:
    print("Invalid value. Try again.")

    









